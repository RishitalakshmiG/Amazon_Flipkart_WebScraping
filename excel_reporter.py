"""
Excel report generation and management
"""
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import EXCEL_OUTPUT_PATH, EXCEL_FILE_NAME

logger = logging.getLogger(__name__)

class ExcelReporter:
    """Generate and update Excel reports"""
    
    def __init__(self, file_path=EXCEL_OUTPUT_PATH):
        """Initialize Excel reporter"""
        self.file_path = file_path
        self.create_workbook()
    
    def create_workbook(self):
        """Create a new workbook with headers"""
        try:
            # Close any existing workbook first
            if hasattr(self, 'wb') and self.wb:
                self.wb.close()
            
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "Product Comparison"
            
            # Set column widths
            self.ws.column_dimensions['A'].width = 25
            self.ws.column_dimensions['B'].width = 15
            self.ws.column_dimensions['C'].width = 15
            self.ws.column_dimensions['D'].width = 12
            self.ws.column_dimensions['E'].width = 12
            self.ws.column_dimensions['F'].width = 15
            self.ws.column_dimensions['G'].width = 12
            self.ws.column_dimensions['H'].width = 12
            self.ws.column_dimensions['I'].width = 20
            self.ws.column_dimensions['J'].width = 30
            self.ws.column_dimensions['K'].width = 30
            
            # Add headers
            headers = [
                "Product Name",
                "Amazon Price",
                "Flipkart Price",
                "Amazon Rating",
                "Flipkart Rating",
                "Amazon Reviews",
                "Flipkart Reviews",
                "Better Deal",
                "Cheaper By %",
                "Amazon URL",
                "Flipkart URL"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = self.ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            self.wb.save(self.file_path)
            self.wb.close()
            logger.info(f"Created workbook at: {self.file_path}")
        except Exception as e:
            logger.error(f"Error creating workbook: {e}")
    
    def add_product_comparison(self, product_data, comparison_result):
        """
        Add product comparison to Excel
        Args:
            product_data: Dictionary with amazon and flipkart data
            comparison_result: Dictionary with comparison results
        """
        try:
            amazon = product_data.get('amazon', {})
            flipkart = product_data.get('flipkart', {})
            
            # Get next row
            next_row = self.ws.max_row + 1
            
            # Style for alternating rows
            row_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid") if next_row % 2 == 0 else None
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            data = [
                amazon.get('product_name', 'N/A'),
                amazon.get('price', 'N/A'),
                flipkart.get('price', 'N/A'),
                amazon.get('rating', 'N/A'),
                flipkart.get('rating', 'N/A'),
                amazon.get('review_count', 0),
                flipkart.get('review_count', 0),
                comparison_result.get('recommendation', 'N/A'),
                comparison_result.get('cheaper_by_percentage', 0),
                amazon.get('url', 'N/A'),
                flipkart.get('url', 'N/A')
            ]
            
            for col, value in enumerate(data, 1):
                cell = self.ws.cell(row=next_row, column=col)
                cell.value = value
                
                # Apply styling
                if row_fill:
                    cell.fill = row_fill
                cell.border = border
                
                # Center align numeric columns
                if col in [2, 3, 4, 5, 6, 7, 8, 9]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            self.wb.save(self.file_path)
            self.wb.close()
            logger.info(f"Added product to Excel: {amazon.get('product_name', 'Unknown')}")
            return True
        except Exception as e:
            logger.error(f"Error adding product to Excel: {e}")
            return False
    
    def update_from_database(self, db_products):
        """
        Update Excel with all products from database
        Args:
            db_products: Dictionary with amazon and flipkart products
        """
        try:
            # Clear existing data (keep headers)
            for row in self.ws.iter_rows(min_row=2, max_row=self.ws.max_row):
                for cell in row:
                    cell.value = None
            
            # Add all products
            amazon_products = db_products.get('amazon', [])
            flipkart_products = db_products.get('flipkart', [])
            
            # Match products by name
            for amazon_product in amazon_products:
                # Find matching Flipkart product
                matching_flipkart = None
                for flipkart_product in flipkart_products:
                    if amazon_product['product_name'][:20].lower() in flipkart_product['product_name'].lower():
                        matching_flipkart = flipkart_product
                        break
                
                if matching_flipkart:
                    product_data = {
                        'amazon': amazon_product,
                        'flipkart': matching_flipkart
                    }
                    comparison_result = self._compare_products(amazon_product, matching_flipkart)
                    self.add_product_comparison(product_data, comparison_result)
            
            self.wb.save(self.file_path)
            self.wb.close()
            logger.info("Updated Excel from database")
            return True
        except Exception as e:
            logger.error(f"Error updating Excel from database: {e}")
            return False
    
    def _compare_products(self, amazon, flipkart):
        """Compare two products and return result"""
        from utils import compare_products
        return compare_products(
            {
                'price': amazon.get('price'),
                'rating': amazon.get('rating', 0),
                'review_count': amazon.get('review_count', 0)
            },
            {
                'price': flipkart.get('price'),
                'rating': flipkart.get('rating', 0),
                'review_count': flipkart.get('review_count', 0)
            }
        )
    
    def get_file_path(self):
        """Get the file path of the Excel file"""
        return self.file_path
#13-12-25